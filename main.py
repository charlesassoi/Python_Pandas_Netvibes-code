
import pandas as pd 



  

#creation du board initial 

filepath="C:/Users/cai5/Documents/Schindler_DS_DT_ESC_IssueManagement_Board_2023_MASTER_20230613 (1) (1) (1) (1) (1) (1) (1) (1).xlsx" 

def board(filepath1,filepath2): 

     

    #date 

        #date1 

    date1=filepath1.split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date1[:4] 

    mois=date1[4:6] 

    jour=date1[6:8] 

    date1=annee+"-"+mois+"-"+jour 

     

        #date2 

    date=filepath2.split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date[:4] 

    mois=date[4:6] 

    jour=date[6:8] 

    date=annee+"-"+mois+"-"+jour 

    date2=date 

     

    #chargement des fichiers 

    df1=pd.DataFrame(pd.read_excel(filepath1,header=4)) 

    df2=pd.DataFrame(pd.read_excel(filepath2,header=4)) 

    df1["New_Priority"]=df1["Priority"].replace(["Medium","Low"],"MediumLow") 

    df2["New_Priority"]=df2["Priority"].replace(["Medium","Low"],"MediumLow") 

    dv = pd.DataFrame(index=['Inflow','Outflow','Open_create','Open_assign','Open_active','Open_review','Open_Sub_total','Open_total_U+H','Open_total','Open_due_date','WIP_DS',"WIP_SCH","SR","CRITSIT","ER",'comments'], columns=['U'+date,'H'+date,'ML'+date]) 

     

    #urgent 

    dw1=df1 

    dw2=df2 

    dw1=dw1[dw1["Priority"]=="Urgent"] 

    dw2=dw2[dw2["Priority"]=="Urgent"] 

    dw22=dw2[dw2["MATURITY"]=="Closed"] 

    dw222=dw2[dw2["MATURITY"]!="Closed"] 

    dw111=dw1[dw1["MATURITY"]=="Closed"] 

    dw11=dw1[dw1["MATURITY"]!="Closed"] 

    mergedStuff = pd.merge(dw1["Reference"], dw222["Reference"], on=['Reference'], how='outer') 

    mergedStuff1 = pd.merge(dw111["Reference"], dw22["Reference"], on=['Reference'], how='outer') 

    for i in range(len(dw1)): 

        mergedStuff=mergedStuff.drop(mergedStuff[mergedStuff["Reference"]==dw1["Reference"].values[i]].index) 

         

        #Outflow1:on retire les closed de la semaine 2 inclus dans les closed de la semaine 1 

    for i in range(len(dw111)): 

        mergedStuff1=mergedStuff1.drop(mergedStuff1[mergedStuff1["Reference"]==dw111["Reference"].values[i]].index) 

         

         

    inflow=len(mergedStuff) 

    outflow=len(mergedStuff1) 

    dv['U'+date].values[dv.index=='Inflow']=inflow 

    dv['U'+date].values[dv.index=='Outflow']=outflow 

    dv['U'+date].values[dv.index=='Open_create']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_assign']=dw222[dw222["MATURITY"]=="Assign"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_active']=dw222[dw222["MATURITY"]=="Active"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_review']=dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_Sub_total']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count()+dw222[dw222["MATURITY"]=="Assign"]["Reference"].count()+dw222[dw222["MATURITY"]=="Active"]["Reference"].count()+dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['U'+date].values[dv.index=='WIP_DS']=dw222[dw222["Corp"]=="DS"]["Reference"].count() 

    dv['U'+date].values[dv.index=='WIP_SCH']=dw222[dw222["Corp"]=="SCH"]["Reference"].count() 

    dv['U'+date].values[dv.index=='SR']=dw222[dw222["Solution\nPath"]=="SR"]["Reference"].count() 

    dv['U'+date].values[dv.index=='ER']=dw222[dw222["Solution\nPath"]=="ER"]["Reference"].count() 

     

    #high 

     

    dw1=df1 

    dw2=df2 

    dw1=dw1[dw1["Priority"]=="High"] 

    dw2=dw2[dw2["Priority"]=="High"] 

    dw22=dw2[dw2["MATURITY"]=="Closed"] 

    dw222=dw2[dw2["MATURITY"]!="Closed"] 

    dw111=dw1[dw1["MATURITY"]=="Closed"] 

    dw11=dw1[dw1["MATURITY"]!="Closed"] 

    mergedStuff = pd.merge(dw1["Reference"], dw222["Reference"], on=['Reference'], how='outer') 

    mergedStuff1 = pd.merge(dw111["Reference"], dw22["Reference"], on=['Reference'], how='outer') 

    for i in range(len(dw1)): 

        mergedStuff=mergedStuff.drop(mergedStuff[mergedStuff["Reference"]==dw1["Reference"].values[i]].index) 

    for i in range(len(dw111)): 

        mergedStuff1=mergedStuff1.drop(mergedStuff1[mergedStuff1["Reference"]==dw111["Reference"].values[i]].index) 

    for i in range(len(mergedStuff1)): 

        dw22=dw22.drop(dw22[dw22["Reference"]==mergedStuff1["Reference"].values[i]].index) 

    inflow=len(mergedStuff) 

    outflow=len(mergedStuff1) 

    dv['H'+date].values[dv.index=='Inflow']=inflow 

    dv['H'+date].values[dv.index=='Outflow']=outflow 

    dv['H'+date].values[dv.index=='Open_create']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_assign']=dw222[dw222["MATURITY"]=="Assign"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_active']=dw222[dw222["MATURITY"]=="Active"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_review']=dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_Sub_total']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count()+dw222[dw222["MATURITY"]=="Assign"]["Reference"].count()+dw222[dw222["MATURITY"]=="Active"]["Reference"].count()+dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['H'+date].values[dv.index=='WIP_DS']=dw222[dw222["Corp"]=="DS"]["Reference"].count() 

    dv['H'+date].values[dv.index=='WIP_SCH']=dw222[dw222["Corp"]=="SCH"]["Reference"].count() 

    dv['H'+date].values[dv.index=='SR']=dw222[dw222["Solution\nPath"]=="SR"]["Reference"].count() 

    dv['H'+date].values[dv.index=='ER']=dw222[dw222["Solution\nPath"]=="ER"]["Reference"].count() 

     

    #ML 

     

    dw1=df1 

    dw2=df2 

    dw1=dw1[dw1["New_Priority"]=="MediumLow"] 

    dw2=dw2[dw2["New_Priority"]=="MediumLow"] 

    dw22=dw2[dw2["MATURITY"]=="Closed"] 

    dw222=dw2[dw2["MATURITY"]!="Closed"] 

    dw111=dw1[dw1["MATURITY"]=="Closed"] 

    dw11=dw1[dw1["MATURITY"]!="Closed"] 

    mergedStuff = pd.merge(dw1["Reference"], dw222["Reference"], on=['Reference'], how='outer') 

    mergedStuff1 = pd.merge(dw111["Reference"], dw22["Reference"], on=['Reference'], how='outer') 

    for i in range(len(dw1)): 

        mergedStuff=mergedStuff.drop(mergedStuff[mergedStuff["Reference"]==dw1["Reference"].values[i]].index) 

    for i in range(len(dw111)): 

        mergedStuff1=mergedStuff1.drop(mergedStuff1[mergedStuff1["Reference"]==dw111["Reference"].values[i]].index) 

    for i in range(len(mergedStuff1)): 

        dw22=dw22.drop(dw22[dw22["Reference"]==mergedStuff1["Reference"].values[i]].index) 

    inflow=len(mergedStuff) 

    outflow=len(mergedStuff1) 

    dv['ML'+date].values[dv.index=='Inflow']=inflow 

    dv['ML'+date].values[dv.index=='Outflow']=outflow 

    dv['ML'+date].values[dv.index=='Open_create']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_assign']=dw222[dw222["MATURITY"]=="Assign"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_active']=dw222[dw222["MATURITY"]=="Active"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_review']=dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_Sub_total']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count()+dw222[dw222["MATURITY"]=="Assign"]["Reference"].count()+dw222[dw222["MATURITY"]=="Active"]["Reference"].count()+dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='WIP_DS']=dw222[dw222["Corp"]=="DS"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='WIP_SCH']=dw222[dw222["Corp"]=="SCH"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='SR']=dw222[dw222["Solution\nPath"]=="SR"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='ER']=dw222[dw222["Solution\nPath"]=="ER"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_total_U+H']=dv['U'+date].values[dv.index=='Open_Sub_total']+dv['H'+date].values[dv.index=='Open_Sub_total'] 

    dv['H'+date].values[dv.index=='Open_total_U+H']=dv['U'+date].values[dv.index=='Open_Sub_total']+dv['H'+date].values[dv.index=='Open_Sub_total'] 

    return [dv,date1,date2] 

  

def BOARD(filepath2): 

     

    #date 

  

    date=filepath2.split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date[:4] 

    mois=date[4:6] 

    jour=date[6:8] 

    date=annee+"-"+mois+"-"+jour 

    date2=date 

     

    #chargement des fichiers 

  

    df2=pd.read_excel(filepath2,header=4) 

    df2["New_Priority"]=df2["Priority"].replace(["Medium","Low"],"MediumLow") 

    dv = pd.DataFrame(index=['Inflow','Outflow','Open_create','Open_assign','Open_active','Open_review','Open_Sub_total','Open_total_U+H','Open_total','Open_due_date','WIP_DS',"WIP_SCH","SR","CRITSIT","ER",'comments'], columns=['U'+date,'H'+date,'ML'+date]) 

     

    #urgent 

  

    dw2=df2 

    dw2=dw2[dw2["Priority"]=="Urgent"] 

    dw22=dw2[dw2["MATURITY"]=="Closed"] 

    dw222=dw2[dw2["MATURITY"]!="Closed"] 

  

     

    inflow=0 

    outflow=0 

    dv['U'+date].values[dv.index=='Inflow']=inflow 

    dv['U'+date].values[dv.index=='Outflow']=outflow 

    dv['U'+date].values[dv.index=='Open_create']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_assign']=dw222[dw222["MATURITY"]=="Assign"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_active']=dw222[dw222["MATURITY"]=="Active"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_review']=dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_Sub_total']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count()+dw222[dw222["MATURITY"]=="Assign"]["Reference"].count()+dw222[dw222["MATURITY"]=="Active"]["Reference"].count()+dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['U'+date].values[dv.index=='WIP_DS']=dw222[dw222["Corp"]=="DS"]["Reference"].count() 

    dv['U'+date].values[dv.index=='WIP_SCH']=dw222[dw222["Corp"]=="SCH"]["Reference"].count() 

    dv['U'+date].values[dv.index=='SR']=dw222[dw222["Solution\nPath"]=="SR"]["Reference"].count() 

    dv['U'+date].values[dv.index=='ER']=dw222[dw222["Solution\nPath"]=="ER"]["Reference"].count() 

     

    #high 

     

    dw2=df2 

    dw2=dw2[dw2["Priority"]=="High"] 

    dw22=dw2[dw2["MATURITY"]=="Closed"] 

    dw222=dw2[dw2["MATURITY"]!="Closed"] 

  

    inflow=0 

    outflow=0 

    dv['H'+date].values[dv.index=='Inflow']=inflow 

    dv['H'+date].values[dv.index=='Outflow']=outflow 

    dv['H'+date].values[dv.index=='Open_create']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_assign']=dw222[dw222["MATURITY"]=="Assign"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_active']=dw222[dw222["MATURITY"]=="Active"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_review']=dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['H'+date].values[dv.index=='Open_Sub_total']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count()+dw222[dw222["MATURITY"]=="Assign"]["Reference"].count()+dw222[dw222["MATURITY"]=="Active"]["Reference"].count()+dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['H'+date].values[dv.index=='WIP_DS']=dw222[dw222["Corp"]=="DS"]["Reference"].count() 

    dv['H'+date].values[dv.index=='WIP_SCH']=dw222[dw222["Corp"]=="SCH"]["Reference"].count() 

    dv['H'+date].values[dv.index=='SR']=dw222[dw222["Solution\nPath"]=="SR"]["Reference"].count() 

    dv['H'+date].values[dv.index=='ER']=dw222[dw222["Solution\nPath"]=="ER"]["Reference"].count() 

     

    #ML 

     

  

    dw2=df2 

    dw2=dw2[dw2["New_Priority"]=="MediumLow"] 

    dw22=dw2[dw2["MATURITY"]=="Closed"] 

    dw222=dw2[dw2["MATURITY"]!="Closed"] 

  

    inflow=0 

    outflow=0 

    dv['ML'+date].values[dv.index=='Inflow']=inflow 

    dv['ML'+date].values[dv.index=='Outflow']=outflow 

    dv['ML'+date].values[dv.index=='Open_create']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_assign']=dw222[dw222["MATURITY"]=="Assign"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_active']=dw222[dw222["MATURITY"]=="Active"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_review']=dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='Open_Sub_total']=dw222[dw222["MATURITY"]=="Create"]["Reference"].count()+dw222[dw222["MATURITY"]=="Assign"]["Reference"].count()+dw222[dw222["MATURITY"]=="Active"]["Reference"].count()+dw222[dw222["Analysis\nSol., Due Date accept"]=="S"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='WIP_DS']=dw222[dw222["Corp"]=="DS"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='WIP_SCH']=dw222[dw222["Corp"]=="SCH"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='SR']=dw222[dw222["Solution\nPath"]=="SR"]["Reference"].count() 

    dv['ML'+date].values[dv.index=='ER']=dw222[dw222["Solution\nPath"]=="ER"]["Reference"].count() 

    dv['U'+date].values[dv.index=='Open_total_U+H']=dv['U'+date].values[dv.index=='Open_Sub_total']+dv['H'+date].values[dv.index=='Open_Sub_total'] 

    dv['H'+date].values[dv.index=='Open_total_U+H']=dv['U'+date].values[dv.index=='Open_Sub_total']+dv['H'+date].values[dv.index=='Open_Sub_total'] 

    return [dv,date2] 

  

  

#dv0=BOARD(filepath)[0] 

  

def concatenation_board(filepathRepertoire): 

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

    #on peut afficher listeFichiers 

  

    dv0=BOARD(filepathRepertoire+"/{}".format(list(listeFichiers)[0]))[0] 

  

    for i in range(len(listeFichiers)-1): 

  

        dv1=board(filepathRepertoire+"/{}".format(list(listeFichiers)[i]),filepathRepertoire+"/{}".format(list(listeFichiers)[i+1]))[0] 

        dz=pd.concat([dv0,dv1],axis=1) 

        dv0=dz 

  

                #enregistrement 

    concatenation_board= dz.to_csv('concatenation_board.csv', index = True) 

    return(dz) 

  

#concatenation_board("C:/Users/cai5/Downloads/Repertoire schindler")    

  

def histogramme(filepathRepertoire): 

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

    #on peut afficher listeFichiers 

    dv0=BOARD((filepathRepertoire+"/{}").format(list(listeFichiers)[0]))[0] 

  

    for i in range(len(listeFichiers)-1): 

     

        dv1=board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[0] 

        dz=pd.concat([dv0,dv1],axis=1) 

        dv0=dz 

  

    du=pd.DataFrame(columns=['New_Urgent_High','Closed_Urgent_High' ,'dates'])    

    df = pd.DataFrame(columns=['New_Urgent_High','Closed_Urgent_High' ,'dates']) 

  

    for i in range(len(listeFichiers)-1): 

        df.loc[i]=[dz['U'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[1]].values[dz.index=='Inflow']+dz['U'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[2]].values[dz.index=='Inflow']+dz['H'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[1]].values[dz.index=='Inflow']+dz['H'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),filepathRepertoire+"/{}".format(list(listeFichiers)[i+1]))[2]].values[dz.index=='Inflow'],dz['U'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[1]].values[dz.index=='Outflow']+dz['U'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[2]].values[dz.index=='Outflow']+dz['H'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[1]].values[dz.index=='Outflow']+dz['H'+board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[2]].values[dz.index=='Outflow'],board((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),(filepathRepertoire+"/{}").format(list(listeFichiers)[i+1]))[2]] 

    df=df.iloc[0::2] 

     

    du=pd.DataFrame(columns=['New_Urgent_High','Closed_Urgent_High' ,'dates']) 

    du1=pd.DataFrame(columns=['New_Urgent_High']) 

    du2=pd.DataFrame(columns=['Closed_Urgent_High']) 

    du3=pd.DataFrame(columns=['dates']) 

  

  

    dx=pd.DataFrame(columns=['New_Urgent_High','Closed_Urgent_High' ,'dates']) 

    for j in range(len(df)): 

        for i in range(df["New_Urgent_High"].values[j][0]): 

            du1.loc[i]=["New_Urgent_High"] 

        for i in range(df["Closed_Urgent_High"].values[j][0]): 

            du2.loc[i]=["Closed_Urgent_High"] 

        for i in range(max(df["New_Urgent_High"].values[j][0],df["Closed_Urgent_High"].values[j][0])): 

            du3.loc[i]=[df["dates"].values[j]] 

        du=pd.concat([du1,du2,du3],axis=1) 

  

        dx=pd.concat([dx,du],axis=0) 

        du=pd.DataFrame(columns=['New_Urgent_High','Closed_Urgent_High' ,'dates']) 

        du1=pd.DataFrame(columns=['New_Urgent_High']) 

        du2=pd.DataFrame(columns=['Closed_Urgent_High']) 

        du3=pd.DataFrame(columns=['dates']) 

  

                #enregistrement 

    histogramme = dx.to_csv('histogramme.csv', index = True) 

    return(histogramme) 

  

  

def courbes(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    dc=pd.DataFrame(columns=["dates_Total_Urgent","dates_Total_Urgent+High"]) 

    dc1=pd.DataFrame(columns=["dates_Total_Urgent"]) 

    dc2=pd.DataFrame(columns=["dates_Total_Urgent+High"]) 

    dcx=pd.DataFrame(columns=["dates_Total_Urgent","dates_Total_Urgent+High"]) 

    dz=concatenation_board(filepathRepertoire)    

  

    for j in range(len(listeFichiers)-1): 

        for i in range(dz["U"+board((filepathRepertoire+"/{}").format(list(listeFichiers)[j]),(filepathRepertoire+"/{}").format(list(listeFichiers)[j+1]))[2]].values[dz.index=='Open_Sub_total'][0]): 

            dc1.loc[i]=board((filepathRepertoire+"/{}").format(list(listeFichiers)[j]),(filepathRepertoire+"/{}").format(list(listeFichiers)[j+1]))[2] 

        for i in range(dz["U"+board((filepathRepertoire+"/{}").format(list(listeFichiers)[j]),(filepathRepertoire+"/{}").format(list(listeFichiers)[j+1]))[2]].values[dz.index=='Open_total_U+H'][0]): 

            dc2.loc[i]=board((filepathRepertoire+"/{}").format(list(listeFichiers)[j]),(filepathRepertoire+"/{}").format(list(listeFichiers)[j+1]))[2] 

        dc=pd.concat([dc1,dc2],axis=1) 

        dcx=pd.concat([dcx,dc],axis=0) 

        dc=pd.DataFrame(columns=["dates_Total_Urgent","dates_Total_Urgent+High"]) 

        dc1=pd.DataFrame(columns=["dates_Total_Urgent"]) 

        dc2=pd.DataFrame(columns=["dates_Total_Urgent+High"]) 

  

            #enregistrement 

    courbes = dcx.to_csv('courbes.csv', index = True) 

    return(courbes) 

  

def TAT_P(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    df=pd.read_excel(filepathRepertoire+"/{}".format(list(listeFichiers)[0]),header=4) 

         

    #date                  

    date=(filepathRepertoire+"/{}".format(list(listeFichiers)[0])).split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date[:4] 

    mois=date[4:6] 

    jour=date[6:8] 

    date=annee+"-"+mois+"-"+jour 

                          

    dvP = pd.DataFrame(index=['P less than 3 month','P 3 to 6 months','P more than 6 months'], columns=[date]) 

    dvP[date].values[dvP.index=='P less than 3 month']=df[df["P\nAge"]=="< 3"]["Reference"].count() 

    dvP[date].values[dvP.index=='P more than 6 months']=df[df["P\nAge"]=="> 6"]["Reference"].count() 

    dvP[date].values[dvP.index=='P 3 to 6 months']=df[df["P\nAge"]=="3 to 6"]["Reference"].count() 

    dTAT_P=dvP 

    for i in range(1,len(listeFichiers)): 

        df=pd.read_excel(filepathRepertoire+"/{}".format(list(listeFichiers)[i]),header=4) 

         

        #date                  

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

                          

        dvP = pd.DataFrame(index=['P less than 3 month','P 3 to 6 months','P more than 6 months'], columns=[date]) 

        dvP[date].values[dvP.index=='P less than 3 month']=df[df["P\nAge"]=="< 3"]["Reference"].count() 

        dvP[date].values[dvP.index=='P more than 6 months']=df[df["P\nAge"]=="> 6"]["Reference"].count() 

        dvP[date].values[dvP.index=='P 3 to 6 months']=df[df["P\nAge"]=="3 to 6"]["Reference"].count() 

        dTAT_P=pd.concat([dTAT_P,dvP],axis=1) 

         

            #enregistrement 

    TAT_P = dTAT_P.to_csv('dTAT_P.csv', index = True) 

    return(TAT_P) 

                                                                       

#TAT_P("C:/Users/cai5/Downloads/Repertoire schindler")  

  

  

def TAT_SR(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    df=pd.read_excel(filepathRepertoire.format(list(listeFichiers)[0]),header=4) 

         

    #date                  

    date=(filepathRepertoire+"/{}".format(list(listeFichiers)[0])).split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date[:4] 

    mois=date[4:6] 

    jour=date[6:8] 

    date=annee+"-"+mois+"-"+jour 

                          

    dvSR = pd.DataFrame(index=['SR less than 3 month','SR 3 to 6 months','SR more than 6 months'], columns=[date]) 

    dvSR[date].values[dvSR.index=='SR less than 3 month']=df[df["SR\nAge"]=="< 3"]["Reference"].count() 

    dvSR[date].values[dvSR.index=='SR more than 6 months']=df[df["SR\nAge"]=="> 6"]["Reference"].count() 

    dvSR[date].values[dvSR.index=='SR 3 to 6 months']=df[df["SR\nAge"]=="3 to 6"]["Reference"].count() 

    dTAT_SR=dvSR 

    for i in range(1,len(listeFichiers)): 

        df=pd.read_excel(filepathRepertoire+"/{}".format(list(listeFichiers)[i]),header=4) 

         

        #date                  

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

                          

        dvSR = pd.DataFrame(index=['SR less than 3 month','SR 3 to 6 months','SR more than 6 months'], columns=[date]) 

        dvSR[date].values[dvSR.index=='SR less than 3 month']=df[df["SR\nAge"]=="< 3"]["Reference"].count() 

        dvSR[date].values[dvSR.index=='SR more than 6 months']=df[df["SR\nAge"]=="> 6"]["Reference"].count() 

        dvSR[date].values[dvSR.index=='SR 3 to 6 months']=df[df["SR\nAge"]=="3 to 6"]["Reference"].count() 

        dTAT_SR=pd.concat([dTAT_SR,dvSR],axis=1) 

         

            #enregistrement 

    TAT_SR = dTAT_SR.to_csv('dTAT_SR.csv', index = True) 

    return(TAT_SR) 

  

#TAT_SR("C:/Users/cai5/Downloads/Repertoire schindler")  

  

def TAT_ER(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    df=pd.read_excel(filepathRepertoire+"/{}".format(list(listeFichiers)[0]),header=4) 

         

    #date                  

    date=(filepathRepertoire+"/{}".format(list(listeFichiers)[0])).split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date[:4] 

    mois=date[4:6] 

    jour=date[6:8] 

    date=annee+"-"+mois+"-"+jour 

                          

    dvER = pd.DataFrame(index=['ER less than 3 month','ER 3 to 6 months','ER more than 6 months'], columns=[date]) 

    dvER[date].values[dvER.index=='ER less than 3 month']=df[df["ER\nAge"]=="< 3"]["Reference"].count() 

    dvER[date].values[dvER.index=='ER more than 6 months']=df[df["ER\nAge"]=="> 6"]["Reference"].count() 

    dvER[date].values[dvER.index=='ER 3 to 6 months']=df[df["ER\nAge"]=="3 to 6"]["Reference"].count() 

    dTAT_ER=dvER 

    for i in range(1,len(listeFichiers)): 

        df=pd.read_excel(filepathRepertoire+"/{}".format(list(listeFichiers)[i]),header=4) 

         

        #date                  

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

                          

        dvER = pd.DataFrame(index=['ER less than 3 month','ER 3 to 6 months','ER more than 6 months'], columns=[date]) 

        dvER[date].values[dvER.index=='ER less than 3 month']=df[df["ER\nAge"]=="< 3"]["Reference"].count() 

        dvER[date].values[dvER.index=='ER more than 6 months']=df[df["ER\nAge"]=="> 6"]["Reference"].count() 

        dvER[date].values[dvER.index=='ER 3 to 6 months']=df[df["ER\nAge"]=="3 to 6"]["Reference"].count() 

        dTAT_ER=pd.concat([dTAT_ER,dvER],axis=1) 

     

        #enregistrement 

    TAT_ER = dTAT_ER.to_csv('dTAT_ER.csv', index = True) 

    return(TAT_ER) 

  

#TAT_ER("C:/Users/cai5/Downloads/Repertoire schindler")  

  

def TAT_Issues(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    df=pd.read_excel((filepathRepertoire+"/{}").format(list(listeFichiers)[0]),header=4) 

         

    #date                  

    date=((filepathRepertoire+"/{}").format(list(listeFichiers)[0])).split("/")[-1].split("_")[-1].split(".")[0] 

    annee=date[:4] 

    mois=date[4:6] 

    jour=date[6:8] 

    date=annee+"-"+mois+"-"+jour 

                          

    dvIssues = pd.DataFrame(index=['Issues less than 3 month','Issues 3 to 6 months','Issues more than 6 months'], columns=[date]) 

    dvIssues[date].values[dvIssues.index=='Issues less than 3 month']=df[df["All Issues\nAge"]=="< 3"]["Reference"].count() 

    dvIssues[date].values[dvIssues.index=='Issues more than 6 months']=df[df["All Issues\nAge"]=="> 6"]["Reference"].count() 

    dvIssues[date].values[dvIssues.index=='Issues 3 to 6 months']=df[df["All Issues\nAge"]=="3 to 6"]["Reference"].count() 

    dTAT_Issues=dvIssues 

    for i in range(1,len(listeFichiers)): 

        df=pd.read_excel((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),header=4) 

         

        #date                  

        date=((filepathRepertoire+"/{}").format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

                          

        dvIssues = pd.DataFrame(index=['Issues less than 3 month','Issues 3 to 6 months','Issues more than 6 months'], columns=[date]) 

        dvIssues[date].values[dvIssues.index=='Issues less than 3 month']=df[df["All Issues\nAge"]=="< 3"]["Reference"].count() 

        dvIssues[date].values[dvIssues.index=='Issues more than 6 months']=df[df["All Issues\nAge"]=="> 6"]["Reference"].count() 

        dvIssues[date].values[dvIssues.index=='Issues 3 to 6 months']=df[df["All Issues\nAge"]=="3 to 6"]["Reference"].count() 

        dTAT_Issues=pd.concat([dTAT_Issues,dvIssues],axis=1) 

         

    #enregistrement 

    TAT_Issues = dTAT_Issues.to_csv('TAT_Issues.csv', index = True) 

     

    return(TAT_Issues) 

  

#TAT_Issues("C:/Users/cai5/Downloads/Repertoire schindler")  

  

def hist_TAT_Issues(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    dm=pd.DataFrame([]) 

    for i in range(len(listeFichiers)): 

        df=pd.read_excel((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),header=4) 

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

  

        dx=pd.DataFrame( columns=["fields",date]) 

        dvIssues = pd.DataFrame(index=['Issues less than 3 month','Issues 3 to 6 months','Issues more than 6 months'], columns=[date]) 

        dvIssues[date].values[dvIssues.index=='Issues less than 3 month']=df[df["All Issues\nAge"]=="< 3"]["Reference"].count() 

        dvIssues[date].values[dvIssues.index=='Issues more than 6 months']=df[df["All Issues\nAge"]=="> 6"]["Reference"].count() 

        dvIssues[date].values[dvIssues.index=='Issues 3 to 6 months']=df[df["All Issues\nAge"]=="3 to 6"]["Reference"].count() 

        for i in range(dvIssues[date].values[dvIssues.index=='Issues less than 3 month'][0]): 

            dx.loc[i]=["less than 3 month",date] 

        for i in range(dvIssues[date].values[dvIssues.index=='Issues less than 3 month'][0],dvIssues[date].values[dvIssues.index=='Issues less than 3 month'][0]+1+dvIssues[date].values[dvIssues.index=='Issues 3 to 6 months'][0]): 

            dx.loc[i]=['3 to 6 months',date] 

        c1=dvIssues[date].values[dvIssues.index=='Issues less than 3 month'][0]+dvIssues[date].values[dvIssues.index=='Issues 3 to 6 months'][0] 

        for i in range(c1,c1+dvIssues[date].values[dvIssues.index=='Issues more than 6 months'][0]+1): 

            dx.loc[i]=['more than 6 months',date] 

        dx.rename(columns={date: 'date'}, inplace=True) 

        dm=pd.concat([dm,dx],axis=0) 

     

    #enregistrement 

    hist_TAT_Issues = dm.to_csv('hist_TAT_Issues.csv', index = True) 

     

    return(hist_TAT_Issues) 

  

#hist_TAT_Issues("C:/Users/cai5/Downloads/Repertoire schindler") 

  

def hist_TAT_P(filepathRepertoire): 

     

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    dm=pd.DataFrame([]) 

    for i in range(len(listeFichiers)): 

        df=pd.read_excel((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),header=4) 

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

  

        dx=pd.DataFrame( columns=["fields",date]) 

        dvP = pd.DataFrame(index=['P less than 3 month','P 3 to 6 months','P more than 6 months'], columns=[date]) 

        dvP[date].values[dvP.index=='P less than 3 month']=df[df["P\nAge"]=="< 3"]["Reference"].count() 

        dvP[date].values[dvP.index=='P more than 6 months']=df[df["P\nAge"]=="> 6"]["Reference"].count() 

        dvP[date].values[dvP.index=='P 3 to 6 months']=df[df["P\nAge"]=="3 to 6"]["Reference"].count() 

        for i in range(dvP[date].values[dvP.index=='P less than 3 month'][0]): 

            dx.loc[i]=["less than 3 month",date] 

        for i in range(dvP[date].values[dvP.index=='P less than 3 month'][0],dvP[date].values[dvP.index=='P less than 3 month'][0]+1+dvP[date].values[dvP.index=='P 3 to 6 months'][0]): 

            dx.loc[i]=['3 to 6 months',date] 

        c1=dvP[date].values[dvP.index=='P less than 3 month'][0]+dvP[date].values[dvP.index=='P 3 to 6 months'][0] 

        for i in range(c1,c1+dvP[date].values[dvP.index=='P more than 6 months'][0]+1): 

            dx.loc[i]=['more than 6 months',date] 

        dx.rename(columns={date: 'date'}, inplace=True) 

        dm=pd.concat([dm,dx],axis=0) 

         

    #enregistrement 

    hist_TAT_P = dm.to_csv('hist_TAT_P.csv', index = True) 

    return(hist_TAT_P) 

  

def hist_TAT_SR(filepathRepertoire): 

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    dm=pd.DataFrame([]) 

    for i in range(len(listeFichiers)): 

        df=pd.read_excel((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),header=4) 

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

  

        dx=pd.DataFrame( columns=["fields",date]) 

        dvSR = pd.DataFrame(index=['SR less than 3 month','SR 3 to 6 months','SR more than 6 months'], columns=[date]) 

        dvSR[date].values[dvSR.index=='SR less than 3 month']=df[df["SR\nAge"]=="< 3"]["Reference"].count() 

        dvSR[date].values[dvSR.index=='SR more than 6 months']=df[df["SR\nAge"]=="> 6"]["Reference"].count() 

        dvSR[date].values[dvSR.index=='SR 3 to 6 months']=df[df["SR\nAge"]=="3 to 6"]["Reference"].count() 

        for i in range(dvSR[date].values[dvSR.index=='SR less than 3 month'][0]): 

            dx.loc[i]=["less than 3 month",date] 

        for i in range(dvSR[date].values[dvSR.index=='SR less than 3 month'][0],dvSR[date].values[dvSR.index=='SR less than 3 month'][0]+1+dvSR[date].values[dvSR.index=='SR 3 to 6 months'][0]): 

            dx.loc[i]=['3 to 6 months',date] 

        c1=dvSR[date].values[dvSR.index=='SR less than 3 month'][0]+dvSR[date].values[dvSR.index=='SR 3 to 6 months'][0] 

        for i in range(c1,c1+dvSR[date].values[dvSR.index=='SR more than 6 months'][0]+1): 

            dx.loc[i]=['more than 6 months',date] 

        dx.rename(columns={date: 'date'}, inplace=True) 

        dm=pd.concat([dm,dx],axis=0) 

         

    #enregistrement 

    hist_TAT_SR = dm.to_csv('hist_TAT_SR.csv', index = True) 

     

    return(hist_TAT_SR) 

  

def hist_TAT_ER(filepathRepertoire): 

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

         

    dm=pd.DataFrame([]) 

    for i in range(len(listeFichiers)): 

        df=pd.read_excel((filepathRepertoire+"/{}").format(list(listeFichiers)[i]),header=4) 

        date=(filepathRepertoire+"/{}".format(list(listeFichiers)[i])).split("/")[-1].split("_")[-1].split(".")[0] 

        annee=date[:4] 

        mois=date[4:6] 

        jour=date[6:8] 

        date=annee+"-"+mois+"-"+jour 

  

        dx=pd.DataFrame( columns=["fields",date]) 

        dvER = pd.DataFrame(index=['ER less than 3 month','ER 3 to 6 months','ER more than 6 months'], columns=[date]) 

        dvER[date].values[dvER.index=='ER less than 3 month']=df[df["ER\nAge"]=="< 3"]["Reference"].count() 

        dvER[date].values[dvER.index=='ER more than 6 months']=df[df["ER\nAge"]=="> 6"]["Reference"].count() 

        dvER[date].values[dvER.index=='ER 3 to 6 months']=df[df["ER\nAge"]=="3 to 6"]["Reference"].count() 

        for i in range(dvER[date].values[dvER.index=='ER less than 3 month'][0]): 

            dx.loc[i]=["less than 3 month",date] 

        for i in range(dvER[date].values[dvER.index=='ER less than 3 month'][0],dvER[date].values[dvER.index=='ER less than 3 month'][0]+1+dvER[date].values[dvER.index=='ER 3 to 6 months'][0]): 

            dx.loc[i]=['3 to 6 months',date] 

        c1=dvER[date].values[dvER.index=='ER less than 3 month'][0]+dvER[date].values[dvER.index=='ER 3 to 6 months'][0] 

        for i in range(c1,c1+dvER[date].values[dvER.index=='ER more than 6 months'][0]+1): 

            dx.loc[i]=['more than 6 months',date] 

        dx.rename(columns={date: 'date'}, inplace=True) 

        dm=pd.concat([dm,dx],axis=0) 

      

    #enregistrement 

    hist_TAT_ER = dm.to_csv('hist_TAT_ER.csv', index = True) 

     

    return(hist_TAT_ER) 

  

#hist_TAT_ER("C:/Users/cai5/Downloads/Repertoire schindler") 

#hist_TAT_SR("C:/Users/cai5/Downloads/Repertoire schindler") 

#hist_TAT_P("C:/Users/cai5/Downloads/Repertoire schindler") 

#TAT_Issues("C:/Users/cai5/Downloads/Repertoire schindler") 

#TAT_ER("C:/Users/cai5/Downloads/Repertoire schindler") 

#TAT_SR("C:/Users/cai5/Downloads/Repertoire schindler") 

#TAT_P("C:/Users/cai5/Downloads/Repertoire schindler") 

#TAT_Issues("C:/Users/cai5/Downloads/Repertoire schindler") 

#hist_TAT_Issues("C:/Users/cai5/Downloads/Repertoire schindler") 

#courbes("C:/Users/cai5/Downloads/Repertoire schindler") 

#histogramme("C:/Users/cai5/Downloads/Repertoire schindler")  

#concatenation_board("C:/Users/cai5/Downloads/Repertoire schindler")    

  

  

#Tableau de bord de pilotage 



def suivi(filepathRepertoire): 

    import pandas as pd 

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

  

     

    df=pd.DataFrame(pd.read_excel(list(listeFichiers)[-1],header=4)) 

    df["New_Priority"]=df["Priority"].replace(["Medium","Low"],"MediumLow") 

    df.rename(columns = {'Ad-Hoc\nSelec':'Ad-Hoc Selec', 'Solution\nPath':'ProjectSRER','Prog. Mgt\nDecision':'Prog. Mgt Decision','Analysis\nSol., Due Date accept':'Analysis_Solution','Urgent Impact\nTarget':'Urgent Impact Target','Urgent Date (if not resolved by that date Plan)	':'Urgent Date (if not resolved by that date Plan)', 'DS Project\nMain contact':'DS Project Main contact', 'SCH\nMain Contact':'SCH Main Contact', 'Category\n3DX, Methodology, Spec. App.':'Category', 'SR\nAge':'SR Age', 'P\nAge':'P Age','ER\nAge':'ER Age','All Issues\nAge':'All Issues Age'}, inplace = True) 

    df=df.replace("FD acc.","FD acc") 

    df["Classif_Solution_Path"]=df["ProjectSRER"] 

    df["Classif_MATURITY"]=df["MATURITY"] 

    df.loc[df.Analysis_Solution=="FD acc",["Classif_Solution_Path","Classif_MATURITY"]]='' 

    df.loc[df.MATURITY=="Review","Classif_Solution_Path"]='' 

    df.loc[df.Analysis_Solution=="FD acc","Classif_MATURITY"]='' 

    df["hyperlink"]="http" 

    from openpyxl import load_workbook 

    # Charger le fichier Excel  

    wb= load_workbook(filename=list(listeFichiers)[-1]) 

    # Sélectionner la feuille de calcul  

    ws=wb["ALL ISSUES"] 

    # Récupérer l'hyperlien de la cellule spécifiée  

    #  cellule contenant l'hyperlien :H1 

    L=[] 

    for i in range(len(df)): 

        hyperlink=ws["H"+str(i+2)].hyperlink 

        try: 

            if type(hyperlink.display)=='str': 

                print(hyperlink) 

        except: 

            L.append(i) 

    for i in range(len(df)): 

        hyperlink=ws["H"+str(i+2)].hyperlink 

        if i not in L: 

            df["hyperlink"].values[i]=hyperlink.display[:53] 

            print(hyperlink.display[:53]) 

        else: 

            df["hyperlink"].values[i]="" 

            print("") 

    SchindlerIIE = df.to_csv('SchindlerIIE.csv', index = True) 

    return(df) 

#suivi("C:/Users/cai5/Downloads/Repertoire schindler") 

  

  

#tableau de partition/Pivot 

  

def partition(filepathRepertoire): 

    import os 

    from os import listdir 

    from os.path import isfile, join 

    fichiers = [f for f in listdir(filepathRepertoire) if isfile(join(filepathRepertoire, f))] 

    from os import walk 

    listeFichiers = [] 

    for (repertoire, sousRepertoires, fichiers) in walk(filepathRepertoire): 

        listeFichiers.extend(fichiers) 

     

    df=pd.DataFrame(pd.read_excel(list(listeFichiers)[-1],header=4)) 

    df["New_Priority"]=df["Priority"].replace(["Medium","Low"],"MediumLow") 

    df.rename(columns = {'Ad-Hoc\nSelec':'Ad-Hoc Selec', 'Solution\nPath':'ProjectSRER','Prog. Mgt\nDecision':'Prog. Mgt Decision','Analysis\nSol., Due Date accept':'Analysis_Solution','Urgent Impact\nTarget':'Urgent Impact Target','Urgent Date (if not resolved by that date Plan)	':'Urgent Date (if not resolved by that date Plan)', 'DS Project\nMain contact':'DS Project Main contact', 'SCH\nMain Contact':'SCH Main Contact', 'Category\n3DX, Methodology, Spec. App.':'Category', 'SR\nAge':'SR Age', 'P\nAge':'P Age','ER\nAge':'ER Age','All Issues\nAge':'All Issues Age'}, inplace = True) 

    df=df.replace("FD acc.","FD acc") 

    df["Classif_Solution_Path"]=df["ProjectSRER"] 

    df["Classif_MATURITY"]=df["MATURITY"] 

    df.loc[df.Analysis_Solution=="FD acc",["Classif_Solution_Path","Classif_MATURITY"]]='' 

    df.loc[df.MATURITY=="Review","Classif_Solution_Path"]='' 

    df.loc[df.Analysis_Solution=="FD acc","Classif_MATURITY"]='' 

    df1 = pd.DataFrame(index=['Project','SR' ,'ER','ER_Candidate','Mixed','FD_accepted','In_Review','SCH_topic','Total','TOTAL'], columns=['Urgent','High','MediumLow','P_Urgent','SR_Urgent','ER_Urgent','P_High','SR_High','ER_High','ML']) 

    #case P_Urgent/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['P_Urgent'].values[df1.index=='In_Review']=dv[dv["Analysis_Solution"]=="S"]["Reference"].count() 

    #Case P_Urgent/FD_accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['P_Urgent'].values[df1.index=='FD_accepted']=dv[dv["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case P_Urgent/Project 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['P_Urgent'].values[df1.index=='Project']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case SR_Urgent/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['SR_Urgent'].values[df1.index=='In_Review']=dv[dv["Analysis_Solution"]=="S"]["Reference"].count()

    #Case SR_Urgent/FD_accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['SR_Urgent'].values[df1.index=='FD_accepted']=dv[dv["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case SR_Urgent/SR 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['SR_Urgent'].values[df1.index=='SR']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case ER_Urgent/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"],dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_Urgent'].values[df1.index=='In_Review']=dv[dv["Analysis_Solution"]=="S"]["Reference"].count() 

    #case ER_Urgent/FD accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"],dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_Urgent'].values[df1.index=='FD_accepted']=dv[dv["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case ER_Urgent/ER 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_Urgent'].values[df1.index=='ER']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #Case ER_Urgent/ERCandidate 

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_Urgent'].values[df1.index=='ER_Candidate']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case P_High/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['P_High'].values[df1.index=='In_Review']=dv[dv["Analysis_Solution"]=="S"]["Reference"].count() 

    #Case P_High/FD_accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['P_High'].values[df1.index=='FD_accepted']=dv[dv["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case P_High/Project 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['P_High'].values[df1.index=='Project']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case SR_High/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['SR_High'].values[df1.index=='In_Review']=dv[dv["Analysis_Solution"]=="S"]["Reference"].count() 

    #Case SR_High/FD_accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['SR_High'].values[df1.index=='FD_accepted']=dv[dv["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case SR_High/SR 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['SR_High'].values[df1.index=='SR']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case ER_High/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"],dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_High'].values[df1.index=='In_Review']=dv[dv["Analysis_Solution"]=="S"]["Reference"].count() 

    #case ER_High/FD accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"],dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_High'].values[df1.index=='FD_accepted']=dv[dv["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case ER_High/ER 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_High'].values[df1.index=='ER']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #Case ER_High/ERCandidate 

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ER_High'].values[df1.index=='ER_Candidate']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case ML/In Review 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dw=dw.drop(dw[dw["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='In_Review']=dw[dw["Analysis_Solution"]=="S"]["Reference"].count() 

    #case ML/FD accepted 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dw=dw.drop(dw[dw["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='FD_accepted']=dw[dw["Analysis_Solution"]=="FD acc"]["Reference"].count() 

    #Case ML/Project 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="P"],dw[dw["ProjectSRER"]=="Complex"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='Project']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #Case ML/SR 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SR"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='SR']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #Case ML/ER 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ER"],dw[dw["ProjectSRER"]=="ER P0"],dw[dw["ProjectSRER"]=="ER P1"],dw[dw["ProjectSRER"]=="ER P2"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='ER']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #Case ML/ERCandidate 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="ERCandidate"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='ER_Candidate']=dv[dv["Analysis_Solution"]=="A"]["Reference"].count() 

    #case ML/Mixed 

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dw=dw.drop(dw[dw["MATURITY"]=="Closed"].index) 

    df1['ML'].values[df1.index=='Mixed']=dw[dw["Analysis_Solution"]=="A"]["Reference"].count() 

    #case Project/Urgent 

    df1['Urgent'].values[df1.index=='Project']=df1['P_Urgent'].values[df1.index=='Project'] 

    #case Project/High 

    df1['High'].values[df1.index=='Project']=df1['P_High'].values[df1.index=='Project'] 

    #case SR/Urgent 

    df1['Urgent'].values[df1.index=='SR']=df1['SR_Urgent'].values[df1.index=='SR'] 

    #case SR/High 

    df1['High'].values[df1.index=='SR']=df1['SR_High'].values[df1.index=='SR'] 

    #case ER/Urgent 

    df1['Urgent'].values[df1.index=='ER']=df1['ER_Urgent'].values[df1.index=='ER'] 

    #case ER/High 

    df1['High'].values[df1.index=='ER']=df1['ER_High'].values[df1.index=='ER'] 

    #case ERCandidate/Urgent 

    df1['Urgent'].values[df1.index=='ER_Candidate']=df1['ER_Urgent'].values[df1.index=='ER_Candidate'] 

    #case ERCandidate/High 

    df1['High'].values[df1.index=='ER_Candidate']=df1['ER_High'].values[df1.index=='ER_Candidate'] 

    #case FD accepted/Urgent 

    df1['Urgent'].values[df1.index=='FD_accepted']=df1['P_Urgent'].values[df1.index=='FD_accepted']+df1['SR_Urgent'].values[df1.index=='FD_accepted']+df1['ER_Urgent'].values[df1.index=='FD_accepted'] 

    #case FD accepted/High 

    df1['High'].values[df1.index=='FD_accepted']=df1['P_High'].values[df1.index=='FD_accepted']+df1['SR_High'].values[df1.index=='FD_accepted']+df1['ER_High'].values[df1.index=='FD_accepted'] 

    #case In review/Urgent 

    df1['Urgent'].values[df1.index=='In_Review']=df1['P_Urgent'].values[df1.index=='In_Review']+df1['SR_Urgent'].values[df1.index=='In_Review']+df1['ER_Urgent'].values[df1.index=='In_Review'] 

    #case In Review/High 

    df1['High'].values[df1.index=='In_Review']=df1['P_High'].values[df1.index=='In_Review']+df1['SR_High'].values[df1.index=='In_Review']+df1['ER_High'].values[df1.index=='In_Review'] 

    #case MediumLow 

    df1["MediumLow"]=df1["ML"] 

    #Case SCH Topic/Urgent 

  

    dw=df 

    dw=dw[dw["New_Priority"]=="Urgent"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SCH"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['Urgent'].values[df1.index=='SCH_topic']=dv[dv["ProjectSRER"]=="SCH"]["Reference"].count() 

    #Case SCH Topic/High 

  

    dw=df 

    dw=dw[dw["New_Priority"]=="High"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SCH"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1['High'].values[df1.index=='SCH_topic']=dv[dv["ProjectSRER"]=="SCH"]["Reference"].count() 

    #Case SCH Topic/MediumLow 

  

    dw=df 

    dw=dw[dw["New_Priority"]=="MediumLow"] 

    dv=pd.concat([dw[dw["ProjectSRER"]=="SCH"]],axis=0) 

    dv=dv.drop(dv[dv["MATURITY"]=="Closed"].index) 

    df1["MediumLow"].values[df1.index=='SCH_topic']=dv[dv["ProjectSRER"]=="SCH"]["Reference"].count() 

    #Case Total/urgent 

    df1["Urgent"].values[df1.index=='Total']=df1["Urgent"].sum() 

  

    #Case Total/High 

    df1["High"].values[df1.index=='Total']=df1["High"].sum() 

  

    #Case Total/MediumLow 

    df1["MediumLow"].values[df1.index=='Total']=df1['ML'].values[df1.index=='Mixed']+df1["MediumLow"].values[df1.index=='SCH_topic']+df1['ML'].values[df1.index=='In_Review']+df1['ML'].values[df1.index=='FD_accepted'] 

    df1=df1.iloc[:,:3] 

    df2=pd.DataFrame(columns=["Priority","fields"]) 

    #Dataframe intermédiaire 

  

    for i in range(df1["Urgent"].values[df1.index=='Project'][0]): 

        df2.loc[i]=["Urgent","Project"] 

    for i in range(df1["Urgent"].values[df1.index=='Project'][0],df1["Urgent"].values[df1.index=='Project'][0]+1+df1["Urgent"].values[df1.index=='SR'][0]): 

        df2.loc[i]=["Urgent","SR"] 

    c1=df1["Urgent"].values[df1.index=='Project'][0]+df1["Urgent"].values[df1.index=='SR'][0] 

    for i in range(c1,c1+df1["Urgent"].values[df1.index=='ER'][0]+1): 

        df2.loc[i]=["Urgent","ER"] 

    c2=c1+df1["Urgent"].values[df1.index=='ER'][0] 

    for i in range(c2,c2+df1["Urgent"].values[df1.index=='ER_Candidate'][0]+1): 

        df2.loc[i]=["Urgent","ER_Candidate"] 

    c3=c2+df1["Urgent"].values[df1.index=='ER_Candidate'][0] 

    for i in range(c3,c3+df1["Urgent"].values[df1.index=='FD_accepted'][0]+1): 

        df2.loc[i]=["Urgent","FD_accepted"] 

    c4=c3+df1["Urgent"].values[df1.index=='FD_accepted'][0] 

    for i in range(c4,c4+df1["Urgent"].values[df1.index=='In_Review'][0]+1): 

        df2.loc[i]=["Urgent","In_Review"] 

    c5=c4+df1["Urgent"].values[df1.index=='In_Review'][0] 

    for i in range(c5,c5+df1["Urgent"].values[df1.index=='SCH_topic'][0]+1): 

        df2.loc[i]=["Urgent","SCH_topic"] 

  

  

  

    for i in range(df1["High"].values[df1.index=='Project'][0]): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","Project"] 

    for i in range(df1["High"].values[df1.index=='Project'][0],df1["High"].values[df1.index=='Project'][0]+1+df1["High"].values[df1.index=='SR'][0]): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","SR"] 

    b1=df1["High"].values[df1.index=='Project'][0]+df1["High"].values[df1.index=='SR'][0] 

    for i in range(b1,b1+df1["High"].values[df1.index=='ER'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","ER"] 

    b2=b1+df1["High"].values[df1.index=='ER'][0] 

    for i in range(b2,b2+df1["High"].values[df1.index=='ER_Candidate'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","ER_Candidate"] 

    b3=b2+df1["High"].values[df1.index=='ER_Candidate'][0] 

    for i in range(b3,b3+df1["High"].values[df1.index=='FD_accepted'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","FD_accepted"] 

    b4=b3+df1["High"].values[df1.index=='FD_accepted'][0] 

    for i in range(b4,b4+df1["High"].values[df1.index=='In_Review'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","In_Review"] 

    b5=b4+df1["High"].values[df1.index=='In_Review'][0] 

    for i in range(b5,b5+df1["High"].values[df1.index=='SCH_topic'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+i]=["High","SCH_topic"] 

  

  

    for i in range(df1["MediumLow"].values[df1.index=='Project'][0]): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","Project"] 

    for i in range(df1["MediumLow"].values[df1.index=='Project'][0],df1["MediumLow"].values[df1.index=='Project'][0]+1+df1["MediumLow"].values[df1.index=='SR'][0]): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","SR"] 

    a1=df1["MediumLow"].values[df1.index=='Project'][0]+df1["MediumLow"].values[df1.index=='SR'][0] 

    for i in range(a1,a1+df1["MediumLow"].values[df1.index=='ER'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","ER"] 

    a2=a1+df1["MediumLow"].values[df1.index=='ER'][0] 

    for i in range(a2,a2+df1["MediumLow"].values[df1.index=='ER_Candidate'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","ER_Candidate"] 

    a3=a2+df1["MediumLow"].values[df1.index=='ER_Candidate'][0] 

    for i in range(a3,a3+df1["MediumLow"].values[df1.index=='FD_accepted'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","FD_accepted"] 

    a4=a3+df1["MediumLow"].values[df1.index=='FD_accepted'][0] 

    for i in range(a4,a4+df1["MediumLow"].values[df1.index=='In_Review'][0]+1): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","In_Review"] 

    a5=a4+df1["MediumLow"].values[df1.index=='In_Review'][0] 

    for i in range(a5,a5+df1["MediumLow"].values[df1.index=='SCH_topic'][0]): 

        df2.loc[df1["Urgent"].values[df1.index=='Total'][0]+df1["High"].values[df1.index=='Total'][0]+i]=["MediumLow","SCH_topic"] 

    SchindlerPivot2 = df2.to_csv('SchindlerPivot2.csv', index = True) 

    return(SchindlerPivot2) 

